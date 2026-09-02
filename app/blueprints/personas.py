"""ABM de Personas (propietarios e inquilinos)."""
from flask import (Blueprint, render_template, redirect, url_for, request,
                   flash, abort, jsonify, make_response)
from flask_login import login_required
from sqlalchemy.exc import IntegrityError

from .. import db
from ..models import Persona, Fiador
from ..utils import normalizar_whatsapp, whatsapp_valido
from ..ui import render_ui

personas_bp = Blueprint("personas", __name__, url_prefix="/personas")


@personas_bp.route("/nueva-rapida", methods=["POST"])
@login_required
def nueva_rapida():
    """Alta rápida (JSON) desde el formulario de contrato: crea la persona con lo mínimo
    (nombre y, opcional, DNI) y la marca como inquilino o propietario según 'rol'."""
    d = request.get_json(silent=True) or {}
    nombre = (d.get("nombre") or "").strip()
    if not nombre:
        return jsonify(ok=False, error="El nombre es obligatorio."), 200
    p = Persona(nombre=nombre, dni=((d.get("dni") or "").strip() or None),
                email=((d.get("email") or "").strip() or None),
                telefono=((d.get("telefono") or "").strip() or None))
    if d.get("rol") == "propietario":
        p.es_propietario = True
    else:
        p.es_inquilino = True
    db.session.add(p)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify(ok=False, error="No se pudo crear (¿DNI ya cargado?)."), 200
    return jsonify(ok=True, id=p.id, nombre=p.nombre, email=p.email, telefono=p.telefono)


@personas_bp.route("/")
@login_required
def listar():
    q = request.args.get("q", "").strip()
    rol = request.args.get("rol", "")
    query = Persona.query
    if q:
        like = f"%{q}%"
        query = query.filter(db.or_(Persona.nombre.ilike(like),
                                    Persona.dni.ilike(like),
                                    Persona.cuit.ilike(like)))
    if rol == "propietario":
        query = query.filter_by(es_propietario=True)
    elif rol == "inquilino":
        query = query.filter_by(es_inquilino=True)

    sort = request.args.get("sort", "")
    direccion = request.args.get("dir", "asc")
    cols = {"nombre": db.func.lower(Persona.nombre), "dni": Persona.dni, "cuit": Persona.cuit,
            "telefono": Persona.telefono, "email": db.func.lower(Persona.email)}
    col = cols.get(sort)
    if col is not None:
        query = query.order_by(col.desc() if direccion == "desc" else col.asc())
    else:
        query = query.order_by(Persona.nombre)
    personas = query.all()
    return render_ui("personas/list.html", personas=personas, q=q, rol=rol)


@personas_bp.route("/exportar-dni-sin-cuit")
@login_required
def exportar_dni_sin_cuit():
    """Excel para completar a mano el CUIT de las personas con DNI cargado y
    sin CUIT.

    No calcula el CUIT acá: el prefijo (20 varón / 27 mujer) depende del sexo
    registrado en AFIP, un dato que este sistema no guarda, y no se puede
    derivar del DNI. Adivinarlo daría un número con el dígito verificador
    "válido" pero potencialmente de otra persona -- inaceptable para algo que
    después se usa para facturar de verdad. En cambio, la planilla trae dos
    columnas para tildar M o F a mano; una fórmula arma el CUIT con el dígito
    verificador correcto para la opción que elijas. Dejar una fila sin tildar
    es señal de "no me importa el CUIT de esa persona"."""
    import re
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    personas = (Persona.query
                .filter(Persona.dni.isnot(None), Persona.dni != "")
                .filter(db.or_(Persona.cuit.is_(None), Persona.cuit == ""))
                .order_by(Persona.nombre).all())

    wb = Workbook()
    ws = wb.active
    ws.title = "DNI sin CUIT"
    ws.append(["Nombre", "DNI", "Rol", "M", "F", "CUIT calculado", "Nota",
              "_PrefijoBase", "_Base10", "_Suma", "_Resto", "_PrefijoFinal", "_Verificador"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    fila = 2
    for p in personas:
        dni_limpio = re.sub(r"\D", "", p.dni or "")
        roles = ", ".join(filter(None, ["Propietario" if p.es_propietario else "",
                                        "Inquilino" if p.es_inquilino else ""])) or "—"
        ws.cell(fila, 1, p.nombre)
        ws.cell(fila, 2, dni_limpio or p.dni)
        ws.cell(fila, 3, roles)
        r = fila
        if len(dni_limpio) in (7, 8):
            dni8 = dni_limpio.zfill(8)
            ws.cell(r, 2, dni8)
            # H: prefijo según M/F. I: base de 10 dígitos para la suma pesada.
            # J/K: suma pesada (coeficientes 5,4,3,2,7,6,5,4,3,2) y resto módulo 11.
            # L: prefijo final -- si el resto da 1, el par (20,27) no tiene
            #    dígito verificador propio y AFIP usa 23 en su lugar.
            # M: dígito verificador final (9 si hubiera sido varón, 4 si mujer,
            #    en el caso del prefijo 23; si no, 11-resto, o 0 si el resto ya es 0).
            ws.cell(r, 8, f'=IF(D{r}<>"","20",IF(E{r}<>"","27",""))')
            ws.cell(r, 9, f'=IF(H{r}="","",H{r}&B{r})')
            ws.cell(r, 10, f'=IF(I{r}="","",SUMPRODUCT(MID(I{r},{{1,2,3,4,5,6,7,8,9,10}},1)*1,'
                          f'{{5,4,3,2,7,6,5,4,3,2}}))')
            ws.cell(r, 11, f'=IF(J{r}="","",MOD(J{r},11))')
            ws.cell(r, 12, f'=IF(K{r}="","",IF(K{r}=1,"23",H{r}))')
            ws.cell(r, 13, f'=IF(K{r}="","",IF(K{r}=0,0,IF(K{r}=1,IF(D{r}<>"",9,4),11-K{r})))')
            ws.cell(r, 6, f'=IF(H{r}="","",L{r}&"-"&B{r}&"-"&M{r})')
            ws.cell(r, 7, f'=IF(AND(D{r}<>"",E{r}<>""),"Tildaste M y F: dejá solo una.","")')
        else:
            ws.cell(r, 7, f"DNI con formato raro ({len(dni_limpio) or '?'} dígitos): "
                         f"revisalo antes de calcular el CUIT a mano.")
        fila += 1

    anchos = {1: 30, 2: 12, 3: 20, 4: 5, 5: 5, 6: 18, 7: 50}
    for col, ancho in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = ancho
    for col in range(8, 14):
        ws.column_dimensions[get_column_letter(col)].hidden = True
    ws.freeze_panes = "A2"
    fill = PatternFill("solid", fgColor="FFF6E5")
    for r in range(2, fila):
        ws.cell(r, 4).fill = fill
        ws.cell(r, 5).fill = fill

    # Segunda hoja, solo informativa: fiadores con DNI. El modelo Fiador no
    # tiene columna de CUIT (no se usa para facturar), así que no hay dónde
    # guardarlo aunque se calcule -- se listan aparte para que no queden
    # escondidos, no para completar.
    fiadores = (Fiador.query.filter(Fiador.dni.isnot(None), Fiador.dni != "")
               .order_by(Fiador.nombre).all())
    if fiadores:
        ws2 = wb.create_sheet("Fiadores (informativo)")
        ws2.append(["Nombre", "DNI", "Nota"])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
        ws2.cell(2, 3, "Los fiadores no tienen campo de CUIT en el sistema "
                       "(no se usa para facturar). Si hace falta guardarlo, avisá y se agrega.")
        for i, f in enumerate(fiadores, start=2):
            ws2.cell(i, 1, f.nombre)
            ws2.cell(i, 2, f.dni)
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 70

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = make_response(buf.read())
    resp.headers["Content-Type"] = ("application/vnd.openxmlformats-officedocument"
                                    ".spreadsheetml.sheet")
    resp.headers["Content-Disposition"] = "attachment; filename=dni_sin_cuit.xlsx"
    return resp


# Islas React desactivadas: se conserva el código y las plantillas por si se
# retoman. Las rutas redirigen a la versión clásica.
@personas_bp.route("/react")
@login_required
def react():
    return redirect(url_for("personas.listar"))


@personas_bp.route("/react/nueva")
@login_required
def react_nueva():
    return redirect(url_for("personas.nueva"))


@personas_bp.route("/react/<int:pid>/editar")
@login_required
def react_editar(pid):
    return redirect(url_for("personas.editar", pid=pid))


def _leer_form(persona):
    persona.nombre = request.form.get("nombre", "").strip()
    persona.dni = request.form.get("dni", "").strip()
    persona.cuit = request.form.get("cuit", "").strip()
    persona.domicilio = request.form.get("domicilio", "").strip()
    persona.localidad = request.form.get("localidad", "").strip()
    persona.telefono = request.form.get("telefono", "").strip()
    persona.email = request.form.get("email", "").strip()
    persona.cond_iva = request.form.get("cond_iva", "").strip()
    persona.es_propietario = bool(request.form.get("es_propietario"))
    persona.es_inquilino = bool(request.form.get("es_inquilino"))
    persona.observaciones = request.form.get("observaciones", "").strip()


def _validar(persona):
    if not persona.nombre:
        return "El nombre es obligatorio."
    if persona.telefono and normalizar_whatsapp(persona.telefono) is None:
        return ("El teléfono no parece válido para WhatsApp. Cargalo con código de área, "
                "ej: 11 2345-6789 o 0221 15-456-7890 (sin dejarlo tan corto).")
    return None


@personas_bp.route("/nueva", methods=["GET", "POST"])
@login_required
def nueva():
    if request.method == "POST":
        persona = Persona()
        _leer_form(persona)
        error = _validar(persona)
        if error:
            flash(error, "error")
            return render_ui("personas/form.html", persona=persona)
        db.session.add(persona)
        db.session.commit()
        flash("Persona creada correctamente.", "ok")
        return redirect(url_for("personas.listar"))
    return render_ui("personas/form.html", persona=Persona())


@personas_bp.route("/<int:pid>/editar", methods=["GET", "POST"])
@login_required
def editar(pid):
    persona = db.session.get(Persona, pid) or abort(404)
    if request.method == "POST":
        _leer_form(persona)
        error = _validar(persona)
        if error:
            flash(error, "error")
            return render_ui("personas/form.html", persona=persona)
        db.session.commit()
        flash("Persona actualizada.", "ok")
        return redirect(url_for("personas.listar"))
    return render_ui("personas/form.html", persona=persona)


@personas_bp.route("/telefonos")
@login_required
def telefonos():
    """Lista los teléfonos que WhatsApp podría rechazar, para corregirlos."""
    personas = (Persona.query
                .filter(Persona.telefono.isnot(None), Persona.telefono != "")
                .order_by(Persona.nombre).all())
    revisar = [p for p in personas if not whatsapp_valido(p.telefono)]
    return render_template("personas/telefonos.html", revisar=revisar,
                           total_con_tel=len(personas))


@personas_bp.route("/<int:pid>/telefono", methods=["POST"])
@login_required
def guardar_telefono(pid):
    """Guarda solo el teléfono de una persona (desde la pantalla de revisión)."""
    persona = db.session.get(Persona, pid) or abort(404)
    d = request.get_json(silent=True) or {}
    persona.telefono = (d.get("telefono") or "").strip()
    db.session.commit()
    return jsonify(ok=True, valido=whatsapp_valido(persona.telefono),
                   telefono=persona.telefono)


@personas_bp.route("/<int:pid>/eliminar", methods=["POST"])
@login_required
def eliminar(pid):
    persona = db.session.get(Persona, pid) or abort(404)
    if persona.inmuebles:
        flash("No se puede eliminar: tiene inmuebles asociados.", "error")
        return redirect(url_for("personas.listar"))
    db.session.delete(persona)
    db.session.commit()
    flash("Persona eliminada.", "ok")
    return redirect(url_for("personas.listar"))
